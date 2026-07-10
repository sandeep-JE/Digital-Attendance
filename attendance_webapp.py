import base64
import binascii
import io
import json
import os
import shutil
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path

import cv2
import face_recognition
import numpy as np
import openpyxl
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)
from PIL import Image
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "attendance_records.xlsx"
FACE_IMAGE_DIR = BASE_DIR / "data" / "employee_faces"

EMPLOYEE_HEADERS = [
    "id",
    "employee_code",
    "full_name",
    "department",
    "status",
    "face_encoding",
    "face_image_path",
    "created_at",
    "updated_at",
]
ATTENDANCE_HEADERS = ["id", "employee_id", "attended_on", "attended_at"]
SETTINGS_HEADERS = ["key", "value"]
USER_HEADERS = [
    "id",
    "username",
    "full_name",
    "role",
    "status",
    "password_hash",
    "security_question",
    "security_answer_hash",
    "created_at",
    "updated_at",
]
DECISION_HEADERS = [
    "id",
    "employee_id",
    "attended_on",
    "duration_seconds",
    "auto_status",
    "final_status",
    "ot_seconds",
    "notes",
    "decided_by",
    "decided_at",
]

DEFAULT_COOLDOWN_SECONDS = 300
PRESENT_SECONDS = 8 * 3600
HALFDAY_SECONDS = 4 * 3600
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"
DEFAULT_ADMIN_QUESTION = "Default admin recovery word"
DEFAULT_ADMIN_ANSWER = "admin"


@dataclass
class AppError(Exception):
    message: str
    status_code: int = 400


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key")
    app.config["WORKBOOK_PATH"] = str(WORKBOOK_PATH)
    app.config["FACE_IMAGE_DIR"] = str(FACE_IMAGE_DIR)

    FACE_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    ensure_workbook_exists()

    @app.context_processor
    def inject_template_context():
        current_user = get_current_user()
        endpoint_titles = {
            "dashboard": "Dashboard",
            "attendance": "Attendance Scanner",
            "employees": "Employees",
            "create_employee": "Add Employee",
            "edit_employee": "Employee Profile",
            "users": "User Management",
            "create_user": "Add User",
            "edit_user": "User Profile",
            "attendance_review": "Final Decisions",
            "settings": "Settings",
        }
        return {
            "current_user": current_user,
            "is_admin": bool(current_user and current_user["role"] == "admin"),
            "current_year": date.today().year,
            "page_meta": {"title": endpoint_titles.get(request.endpoint, "Facial Attendance")},
        }

    @app.errorhandler(AppError)
    def handle_app_error(error):
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": error.message}), error.status_code
        flash(error.message, "error")
        if request.endpoint in {"login", "forgot_password"}:
            return redirect(url_for(request.endpoint))
        return redirect(request.referrer or url_for("dashboard"))

    @app.errorhandler(Exception)
    def handle_unexpected_error(error):
        if isinstance(error, AppError):
            return handle_app_error(error)
        if isinstance(error, HTTPException):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": error.description}), error.code
            return (
                render_template(
                    "error.html",
                    error_title=error.name,
                    error_message=error.description,
                ),
                error.code,
            )

        app.logger.exception("Unhandled application error")
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "An unexpected error occurred. Please try again."}), 500
        return render_template(
            "error.html",
            error_title="Something went wrong",
            error_message=(
                "The request could not be completed. If the Excel workbook is open in another "
                "program, close it and try again."
            ),
        ), 500

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if get_current_user():
            return redirect(url_for("dashboard"))

        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""

            user = find_user_by_username(username)
            if not user or user["status"] != "active":
                raise AppError("Invalid username or inactive user account.")
            if not check_password_hash(user["password_hash"], password):
                raise AppError("Invalid username or password.")

            session["user_id"] = user["id"]
            flash(f"Welcome back, {user['full_name']}.", "success")
            return redirect(url_for("dashboard"))

        return render_template("login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("You have been logged out.", "success")
        return redirect(url_for("login"))

    @app.route("/forgot-password", methods=["GET", "POST"])
    def forgot_password():
        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            security_answer = (request.form.get("security_answer") or "").strip()
            new_password = request.form.get("new_password") or ""
            confirm_password = request.form.get("confirm_password") or ""

            user = find_user_by_username(username)
            if not user:
                raise AppError("User not found.")
            if user["status"] != "active":
                raise AppError("Only active users can reset passwords.")
            if not check_password_hash(user["security_answer_hash"], security_answer.lower()):
                raise AppError("Security answer did not match.")

            validate_new_password(new_password, confirm_password)
            updated_user = {
                **user,
                "password_hash": generate_password_hash(new_password),
                "updated_at": now_iso(),
            }
            update_user(updated_user)
            flash("Password reset successful. Please log in with the new password.", "success")
            return redirect(url_for("login"))

        return render_template("forgot_password.html")

    @app.route("/")
    @login_required
    def dashboard():
        employees = list_employees()
        active_employees = [row for row in employees if row["status"] == "active"]
        attendance_rows = list_attendance()
        decisions = list_decisions()
        today = date.today().isoformat()
        day_summaries = build_day_summaries(today, active_employees, attendance_rows, decisions)

        stats = {
            "employees": len(employees),
            "active_employees": len(active_employees),
            "registered_faces": sum(1 for row in employees if row["face_encoding"]),
            "today_present": sum(1 for row in day_summaries if row["resolved_status"] == "present"),
            "today_halfday": sum(1 for row in day_summaries if row["resolved_status"] == "halfday"),
            "today_pending": sum(1 for row in day_summaries if row["resolved_status"] == "pending"),
        }

        employee_lookup = {row["id"]: row for row in employees}
        recent_attendance = []
        for row in sorted(attendance_rows, key=lambda item: item["attended_at"], reverse=True)[:10]:
            employee = employee_lookup.get(row["employee_id"])
            if employee is None:
                continue
            recent_attendance.append(
                {
                    "id": row["id"],
                    "attended_at": row["attended_at"],
                    "full_name": employee["full_name"],
                    "employee_code": employee["employee_code"],
                    "department": employee["department"],
                }
            )

        return render_template(
            "dashboard.html",
            stats=stats,
            recent_attendance=recent_attendance,
            today_summaries=day_summaries,
        )

    @app.route("/employees")
    @admin_required
    def employees():
        rows = sorted(list_employees(), key=lambda item: item["full_name"].lower())
        return render_template("employees.html", employees=rows)

    @app.route("/employees/new", methods=["GET", "POST"])
    @admin_required
    def create_employee():
        if request.method == "POST":
            payload = validate_employee_form(request.form)
            employees_data = list_employees()

            if find_employee_by_code(payload["employee_code"], employees_data):
                raise AppError("Employee code already exists. Please use a unique code.")

            new_employee = {
                "id": next_id(employees_data),
                "employee_code": payload["employee_code"],
                "full_name": payload["full_name"],
                "department": payload["department"],
                "status": payload["status"],
                "face_encoding": "",
                "face_image_path": "",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
            append_employee(new_employee)
            flash(
                "Employee details saved. Start the camera and capture the face to finish registration.",
                "success",
            )
            return redirect(
                url_for(
                    "edit_employee",
                    employee_id=new_employee["id"],
                    setup="face",
                    _anchor="face-registration",
                )
            )

        return render_template("employee_form.html", employee=None)

    @app.route("/employees/<int:employee_id>/edit", methods=["GET", "POST"])
    @admin_required
    def edit_employee(employee_id: int):
        employees_data = list_employees()
        employee = get_employee_or_404(employee_id, employees_data)

        if request.method == "POST":
            payload = validate_employee_form(request.form)
            existing = find_employee_by_code(payload["employee_code"], employees_data)
            if existing and existing["id"] != employee_id:
                raise AppError("Employee code already exists. Please use a unique code.")

            updated_employee = {
                **employee,
                "employee_code": payload["employee_code"],
                "full_name": payload["full_name"],
                "department": payload["department"],
                "status": payload["status"],
                "updated_at": now_iso(),
            }
            update_employee(updated_employee)
            flash("Employee updated successfully.", "success")
            return redirect(url_for("edit_employee", employee_id=employee_id))

        attendance_rows = list_attendance()
        decisions = list_decisions()
        employee_summaries = build_employee_daily_summaries(employee_id, attendance_rows, decisions)
        return render_template(
            "employee_form.html",
            employee=employee,
            attendance_rows=employee_summaries[:15],
        )

    @app.post("/employees/<int:employee_id>/delete")
    @admin_required
    def delete_employee(employee_id: int):
        employees_data = list_employees()
        employee = get_employee_or_404(employee_id, employees_data)

        if employee["face_image_path"]:
            image_path = BASE_DIR / employee["face_image_path"]
            if image_path.exists():
                image_path.unlink()

        delete_employee_row(employee_id)
        delete_employee_attendance(employee_id)
        delete_decisions_for_employee(employee_id)
        flash("Employee deleted successfully.", "success")
        return redirect(url_for("employees"))

    @app.post("/employees/<int:employee_id>/toggle-status")
    @admin_required
    def toggle_employee_status(employee_id: int):
        employees_data = list_employees()
        employee = get_employee_or_404(employee_id, employees_data)
        new_status = "inactive" if employee["status"] == "active" else "active"

        updated_employee = {**employee, "status": new_status, "updated_at": now_iso()}
        update_employee(updated_employee)
        flash(f"{employee['full_name']} is now {new_status}.", "success")
        return redirect(url_for("employees"))

    @app.route("/users")
    @admin_required
    def users():
        rows = sorted(list_users(), key=lambda item: item["full_name"].lower())
        return render_template("users.html", users=rows)

    @app.route("/users/new", methods=["GET", "POST"])
    @admin_required
    def create_user():
        if request.method == "POST":
            payload = validate_user_form(request.form, is_new=True)
            users_data = list_users()
            if find_user_by_username(payload["username"], users_data):
                raise AppError("Username already exists. Please choose another username.")

            user = {
                "id": next_id(users_data),
                "username": payload["username"],
                "full_name": payload["full_name"],
                "role": payload["role"],
                "status": payload["status"],
                "password_hash": generate_password_hash(payload["password"]),
                "security_question": payload["security_question"],
                "security_answer_hash": generate_password_hash(payload["security_answer"].lower()),
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
            append_user(user)
            flash("User created successfully.", "success")
            return redirect(url_for("users"))

        return render_template("user_form.html", user=None)

    @app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
    @admin_required
    def edit_user(user_id: int):
        users_data = list_users()
        user = get_user_or_404(user_id, users_data)

        if request.method == "POST":
            payload = validate_user_form(request.form, is_new=False)
            existing = find_user_by_username(payload["username"], users_data)
            if existing and existing["id"] != user_id:
                raise AppError("Username already exists. Please choose another username.")

            updated_user = {
                **user,
                "username": payload["username"],
                "full_name": payload["full_name"],
                "role": payload["role"],
                "status": payload["status"],
                "security_question": payload["security_question"],
                "updated_at": now_iso(),
            }

            if payload["security_answer"]:
                updated_user["security_answer_hash"] = generate_password_hash(
                    payload["security_answer"].lower()
                )
            if payload["password"]:
                validate_new_password(payload["password"], payload["confirm_password"])
                updated_user["password_hash"] = generate_password_hash(payload["password"])

            guard_admin_account_changes(user_id, updated_user["role"], updated_user["status"])
            update_user(updated_user)

            if session.get("user_id") == user_id and updated_user["status"] != "active":
                session.clear()
                flash("Your user account was deactivated.", "error")
                return redirect(url_for("login"))

            flash("User updated successfully.", "success")
            return redirect(url_for("edit_user", user_id=user_id))

        return render_template("user_form.html", user=user)

    @app.post("/users/<int:user_id>/delete")
    @admin_required
    def delete_user(user_id: int):
        if session.get("user_id") == user_id:
            raise AppError("You cannot delete the currently logged-in account.")
        user = get_user_or_404(user_id)
        guard_admin_account_changes(user_id, None, "deleted")
        delete_user_row(user_id)
        flash("User deleted successfully.", "success")
        return redirect(url_for("users"))

    @app.post("/users/<int:user_id>/toggle-status")
    @admin_required
    def toggle_user_status(user_id: int):
        user = get_user_or_404(user_id)
        new_status = "inactive" if user["status"] == "active" else "active"
        guard_admin_account_changes(user_id, user["role"], new_status)
        updated_user = {**user, "status": new_status, "updated_at": now_iso()}
        update_user(updated_user)
        flash(f"{user['full_name']} is now {new_status}.", "success")
        return redirect(url_for("users"))

    @app.route("/attendance")
    @login_required
    def attendance():
        employees_data = [
            row
            for row in list_employees()
            if row["status"] == "active" and row["face_encoding"]
        ]
        employees_data.sort(key=lambda item: item["full_name"].lower())
        settings = get_settings()
        today_summary = build_today_summaries(
            [row for row in list_employees() if row["status"] == "active"],
            list_attendance(),
            list_decisions(),
        )
        return render_template(
            "attendance.html",
            employees=employees_data,
            settings=settings,
            today_summary=today_summary,
        )

    @app.route("/attendance/review")
    @admin_required
    def attendance_review():
        selected_date = (request.args.get("date") or date.today().isoformat()).strip()
        employees_data = [row for row in list_employees() if row["status"] == "active"]
        review_rows = [
            row
            for row in build_day_summaries(
                selected_date, employees_data, list_attendance(), list_decisions()
            )
            if row["auto_status"] == "review"
        ]
        return render_template(
            "attendance_review.html",
            review_rows=review_rows,
            selected_date=selected_date,
        )

    @app.post("/attendance/review/save")
    @admin_required
    def save_attendance_review():
        employee_id = parse_required_int(request.form.get("employee_id"), "Employee")
        attended_on = (request.form.get("attended_on") or "").strip()
        final_status = (request.form.get("final_status") or "").strip().lower()
        notes = (request.form.get("notes") or "").strip()

        if final_status not in {"present", "halfday", "absent"}:
            raise AppError("Final attendance status must be present, halfday, or absent.")

        employee = get_employee_or_404(employee_id)
        summary = build_employee_day_summary(
            employee, attended_on, list_attendance(), find_decision(employee_id, attended_on)
        )
        current_user = require_admin_user()

        upsert_decision(
            {
                "employee_id": employee_id,
                "attended_on": attended_on,
                "duration_seconds": summary["duration_seconds"],
                "auto_status": summary["auto_status"],
                "final_status": final_status,
                "ot_seconds": summary["ot_seconds"],
                "notes": notes,
                "decided_by": current_user["username"],
                "decided_at": now_iso(),
            }
        )
        flash(f"Final attendance decision saved for {employee['full_name']}.", "success")
        return redirect(url_for("attendance_review", date=attended_on))

    @app.route("/settings", methods=["GET", "POST"])
    @admin_required
    def settings():
        current_settings = get_settings()

        if request.method == "POST":
            cooldown_seconds = parse_cooldown_seconds(request.form.get("cooldown_seconds"))
            set_setting("cooldown_seconds", str(cooldown_seconds))
            flash("Settings updated successfully.", "success")
            return redirect(url_for("settings"))

        return render_template("settings.html", settings=current_settings)

    @app.route("/face-images/<path:filename>")
    @login_required
    def face_image(filename: str):
        return send_from_directory(Path(app.config["FACE_IMAGE_DIR"]), filename)

    @app.post("/api/employees/<int:employee_id>/register-face")
    @admin_required
    def register_face(employee_id: int):
        employee = get_employee_or_404(employee_id)
        payload = request.get_json(silent=True) or {}
        frame = decode_data_url_image(payload.get("image"))
        encoding = extract_single_face_encoding(frame)

        relative_path = save_employee_face(employee_id, employee["employee_code"], frame)
        updated_employee = {
            **employee,
            "face_encoding": serialize_encoding(encoding),
            "face_image_path": relative_path,
            "updated_at": now_iso(),
        }
        update_employee(updated_employee)

        return jsonify(
            {
                "ok": True,
                "message": f"Face registered successfully for {employee['full_name']}.",
                "image_path": relative_path,
            }
        )

    @app.post("/api/attendance/scan")
    @login_required
    def scan_attendance():
        payload = request.get_json(silent=True) or {}
        frame = decode_data_url_image(payload.get("image"))
        target_encoding = extract_single_face_encoding(frame)

        employees_data = [row for row in list_employees() if row["face_encoding"]]
        if not employees_data:
            raise AppError("No registered faces found. Register at least one employee first.")

        active_employees = [row for row in employees_data if row["status"] == "active"]
        if not active_employees:
            raise AppError("No active employees with registered faces are available.")

        encodings = [deserialize_encoding(row["face_encoding"]) for row in active_employees]
        distances = face_recognition.face_distance(encodings, target_encoding)
        best_index = int(np.argmin(distances))
        best_distance = float(distances[best_index])

        if best_distance > 0.45:
            raise AppError("Face not recognized. Please try again with better lighting and framing.")

        employee = active_employees[best_index]
        attendance_rows = list_attendance()
        cooldown_seconds = get_settings()["cooldown_seconds"]
        now = datetime.now()
        latest_scan = get_latest_scan(employee["id"], attendance_rows)

        if latest_scan is not None:
            latest_time = parse_iso_datetime(latest_scan["attended_at"])
            elapsed = (now - latest_time).total_seconds()
            if elapsed < cooldown_seconds:
                remaining_seconds = int(cooldown_seconds - elapsed)
                raise AppError(
                    f"{employee['full_name']} is in cooldown. Try again in {format_seconds(remaining_seconds)}."
                )

        delete_decision_for_day(employee["id"], now.date().isoformat())
        append_attendance(
            {
                "id": next_id(attendance_rows),
                "employee_id": employee["id"],
                "attended_on": now.date().isoformat(),
                "attended_at": now.isoformat(timespec="seconds"),
            }
        )

        summary = build_employee_day_summary(
            employee,
            now.date().isoformat(),
            list_attendance(),
            find_decision(employee["id"], now.date().isoformat()),
        )
        return jsonify(
            {
                "ok": True,
                "message": f"Timestamp recorded for {employee['full_name']}.",
                "employee": {
                    "id": employee["id"],
                    "full_name": employee["full_name"],
                    "employee_code": employee["employee_code"],
                    "department": employee["department"],
                },
                "attended_at": now.isoformat(timespec="seconds"),
                "day_status": summary["resolved_status"],
                "day_duration": summary["duration_text"],
                "ot_text": summary["ot_text"],
            }
        )

    return app


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def workbook_path():
    return WORKBOOK_PATH


def ensure_workbook_exists():
    if workbook_path().exists():
        try:
            workbook = load_workbook_file()
        except AppError as exc:
            if not is_workbook_corruption_error(exc):
                raise
            backup_corrupted_workbook()
            create_new_workbook()
            workbook = load_workbook_file()
        try:
            dirty = False
            dirty = ensure_sheet(workbook, "Employees", EMPLOYEE_HEADERS) or dirty
            dirty = ensure_sheet(workbook, "Attendance", ATTENDANCE_HEADERS) or dirty
            dirty = ensure_sheet(workbook, "Settings", SETTINGS_HEADERS) or dirty
            dirty = ensure_sheet(workbook, "Users", USER_HEADERS) or dirty
            dirty = ensure_sheet(workbook, "Decisions", DECISION_HEADERS) or dirty
            dirty = ensure_default_setting(workbook, "cooldown_seconds", str(DEFAULT_COOLDOWN_SECONDS)) or dirty
            dirty = ensure_default_admin_user(workbook) or dirty
            if dirty:
                save_workbook(workbook)
        finally:
            workbook.close()
        return

    create_new_workbook()


def create_new_workbook():
    workbook = openpyxl.Workbook()
    workbook.active.title = "Employees"
    workbook["Employees"].append(EMPLOYEE_HEADERS)
    workbook.create_sheet("Attendance").append(ATTENDANCE_HEADERS)
    workbook.create_sheet("Settings").append(SETTINGS_HEADERS)
    workbook["Settings"].append(["cooldown_seconds", str(DEFAULT_COOLDOWN_SECONDS)])
    workbook.create_sheet("Users").append(USER_HEADERS)
    workbook.create_sheet("Decisions").append(DECISION_HEADERS)
    ensure_default_admin_user(workbook)
    save_workbook(workbook)


def ensure_sheet(workbook, sheet_name, headers):
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name).append(headers)
        return True

    sheet = workbook[sheet_name]
    existing_headers = [sheet.cell(row=1, column=index + 1).value for index in range(len(headers))]
    if existing_headers != headers:
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)
        return True
    return False


def ensure_default_setting(workbook, key, value):
    sheet = workbook["Settings"]
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == key:
            return False
    sheet.append([key, value])
    return True


def ensure_default_admin_user(workbook):
    sheet = workbook["Users"]
    for row_index in range(2, sheet.max_row + 1):
        username = normalize_value(sheet.cell(row=row_index, column=2).value)
        if username.lower() == DEFAULT_ADMIN_USERNAME:
            return False

    timestamp = now_iso()
    sheet.append(
        [
            1,
            DEFAULT_ADMIN_USERNAME,
            "System Admin",
            "admin",
            "active",
            generate_password_hash(DEFAULT_ADMIN_PASSWORD),
            DEFAULT_ADMIN_QUESTION,
            generate_password_hash(DEFAULT_ADMIN_ANSWER),
            timestamp,
            timestamp,
        ]
    )
    return True


def load_workbook():
    ensure_workbook_exists()
    return load_workbook_file()


def load_workbook_file():
    try:
        return openpyxl.load_workbook(workbook_path())
    except PermissionError as exc:
        raise AppError(
            "The Excel attendance file is currently locked. Please close 'attendance_records.xlsx' in Excel and try again.",
            status_code=409,
        ) from exc
    except OSError as exc:
        raise AppError("The attendance workbook could not be opened. Please verify the Excel file is available.") from exc
    except zipfile.BadZipFile as exc:
        raise AppError("The attendance workbook is corrupted and could not be opened.") from exc


def save_workbook(workbook):
    try:
        workbook.save(workbook_path())
    except PermissionError as exc:
        raise AppError(
            "The Excel attendance file is currently open in another program. Close 'attendance_records.xlsx' and try again.",
            status_code=409,
        ) from exc
    except OSError as exc:
        raise AppError("The attendance workbook could not be saved. Please check file access and try again.") from exc


def is_workbook_corruption_error(error):
    return "corrupted" in error.message.lower()


def backup_corrupted_workbook():
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_path = workbook_path().with_name(f"attendance_records.corrupt.{timestamp}.xlsx")
    shutil.move(str(workbook_path()), str(backup_path))


def normalize_value(value):
    return "" if value is None else value


def sheet_to_records(sheet, headers):
    records = []
    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row_values):
            continue
        record = {header: normalize_value(value) for header, value in zip(headers, row_values)}
        for integer_key in {"id", "employee_id"}:
            if integer_key in record and record[integer_key] != "":
                record[integer_key] = int(record[integer_key])
        records.append(record)
    return records


def next_id(records):
    if not records:
        return 1
    return max(record["id"] for record in records) + 1


def list_employees():
    workbook = load_workbook()
    try:
        return sheet_to_records(workbook["Employees"], EMPLOYEE_HEADERS)
    finally:
        workbook.close()


def list_attendance():
    workbook = load_workbook()
    try:
        return sheet_to_records(workbook["Attendance"], ATTENDANCE_HEADERS)
    finally:
        workbook.close()


def list_users():
    workbook = load_workbook()
    try:
        return sheet_to_records(workbook["Users"], USER_HEADERS)
    finally:
        workbook.close()


def list_decisions():
    workbook = load_workbook()
    try:
        return sheet_to_records(workbook["Decisions"], DECISION_HEADERS)
    finally:
        workbook.close()


def get_settings():
    workbook = load_workbook()
    try:
        settings = {"cooldown_seconds": DEFAULT_COOLDOWN_SECONDS}
        sheet = workbook["Settings"]
        for row_index in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row_index, column=1).value
            value = sheet.cell(row=row_index, column=2).value
            if key:
                settings[key] = value
        settings["cooldown_seconds"] = int(settings.get("cooldown_seconds", DEFAULT_COOLDOWN_SECONDS))
        settings["cooldown_text"] = format_seconds(settings["cooldown_seconds"])
        return settings
    finally:
        workbook.close()


def set_setting(key, value):
    workbook = load_workbook()
    try:
        sheet = workbook["Settings"]
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_index, column=1).value == key:
                sheet.cell(row=row_index, column=2, value=value)
                save_workbook(workbook)
                return
        sheet.append([key, value])
        save_workbook(workbook)
    finally:
        workbook.close()


def append_employee(employee):
    workbook = load_workbook()
    try:
        workbook["Employees"].append([employee[header] for header in EMPLOYEE_HEADERS])
        save_workbook(workbook)
    finally:
        workbook.close()


def update_employee(employee):
    workbook = load_workbook()
    try:
        update_row_by_id(workbook["Employees"], EMPLOYEE_HEADERS, employee)
        save_workbook(workbook)
    finally:
        workbook.close()


def delete_employee_row(employee_id):
    workbook = load_workbook()
    try:
        delete_row_by_id(workbook["Employees"], employee_id)
        save_workbook(workbook)
    finally:
        workbook.close()


def append_user(user):
    workbook = load_workbook()
    try:
        workbook["Users"].append([user[header] for header in USER_HEADERS])
        save_workbook(workbook)
    finally:
        workbook.close()


def update_user(user):
    workbook = load_workbook()
    try:
        update_row_by_id(workbook["Users"], USER_HEADERS, user)
        save_workbook(workbook)
    finally:
        workbook.close()


def delete_user_row(user_id):
    workbook = load_workbook()
    try:
        delete_row_by_id(workbook["Users"], user_id)
        save_workbook(workbook)
    finally:
        workbook.close()


def append_attendance(record):
    workbook = load_workbook()
    try:
        workbook["Attendance"].append([record[header] for header in ATTENDANCE_HEADERS])
        save_workbook(workbook)
    finally:
        workbook.close()


def delete_employee_attendance(employee_id):
    workbook = load_workbook()
    try:
        sheet = workbook["Attendance"]
        row_index = 2
        while row_index <= sheet.max_row:
            if sheet.cell(row=row_index, column=2).value == employee_id:
                sheet.delete_rows(row_index, 1)
            else:
                row_index += 1
        save_workbook(workbook)
    finally:
        workbook.close()


def upsert_decision(decision):
    decisions = list_decisions()
    existing = next(
        (
            row
            for row in decisions
            if row["employee_id"] == decision["employee_id"] and row["attended_on"] == decision["attended_on"]
        ),
        None,
    )

    workbook = load_workbook()
    try:
        sheet = workbook["Decisions"]
        if existing:
            merged = {"id": existing["id"], **decision}
            update_row_by_id(sheet, DECISION_HEADERS, merged)
        else:
            merged = {"id": next_id(decisions), **decision}
            sheet.append([merged[header] for header in DECISION_HEADERS])
        save_workbook(workbook)
    finally:
        workbook.close()


def delete_decisions_for_employee(employee_id):
    workbook = load_workbook()
    try:
        sheet = workbook["Decisions"]
        row_index = 2
        while row_index <= sheet.max_row:
            if sheet.cell(row=row_index, column=2).value == employee_id:
                sheet.delete_rows(row_index, 1)
            else:
                row_index += 1
        save_workbook(workbook)
    finally:
        workbook.close()


def delete_decision_for_day(employee_id, attended_on):
    workbook = load_workbook()
    try:
        sheet = workbook["Decisions"]
        row_index = 2
        dirty = False
        while row_index <= sheet.max_row:
            if (
                sheet.cell(row=row_index, column=2).value == employee_id
                and sheet.cell(row=row_index, column=3).value == attended_on
            ):
                sheet.delete_rows(row_index, 1)
                dirty = True
            else:
                row_index += 1
        if dirty:
            save_workbook(workbook)
    finally:
        workbook.close()


def update_row_by_id(sheet, headers, record):
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == record["id"]:
            for column_index, header in enumerate(headers, start=1):
                sheet.cell(row=row_index, column=column_index, value=record[header])
            return
    raise AppError("Requested record was not found.", status_code=404)


def delete_row_by_id(sheet, record_id):
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == record_id:
            sheet.delete_rows(row_index, 1)
            return
    raise AppError("Requested record was not found.", status_code=404)


def find_employee_by_code(employee_code, employees_data=None):
    employees_data = employees_data or list_employees()
    for employee in employees_data:
        if employee["employee_code"].lower() == employee_code.lower():
            return employee
    return None


def find_user_by_username(username, users_data=None):
    users_data = users_data or list_users()
    for user in users_data:
        if user["username"].lower() == username.lower():
            return user
    return None


def find_user_by_id(user_id, users_data=None):
    users_data = users_data or list_users()
    for user in users_data:
        if user["id"] == user_id:
            return user
    return None


def find_decision(employee_id, attended_on, decisions=None):
    decisions = decisions or list_decisions()
    for decision in decisions:
        if decision["employee_id"] == employee_id and decision["attended_on"] == attended_on:
            return decision
    return None


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    user = find_user_by_id(user_id)
    if not user or user["status"] != "active":
        session.clear()
        return None
    return user


def require_admin_user():
    user = get_current_user()
    if not user or user["role"] != "admin":
        raise AppError("Administrator access is required.", status_code=403)
    return user


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Login required."}), 401
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapper


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Login required."}), 401
            return redirect(url_for("login"))
        if user["role"] != "admin":
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Administrator access is required."}), 403
            raise AppError("Administrator access is required.", status_code=403)
        return view_func(*args, **kwargs)

    return wrapper


def get_employee_or_404(employee_id, employees_data=None):
    employees_data = employees_data or list_employees()
    for employee in employees_data:
        if employee["id"] == employee_id:
            return employee
    raise AppError("Employee not found.", status_code=404)


def get_user_or_404(user_id, users_data=None):
    users_data = users_data or list_users()
    for user in users_data:
        if user["id"] == user_id:
            return user
    raise AppError("User not found.", status_code=404)


def guard_admin_account_changes(user_id, new_role, new_status):
    users_data = list_users()
    admin_users = [
        row
        for row in users_data
        if row["role"] == "admin" and row["status"] == "active" and row["id"] != user_id
    ]
    current = get_user_or_404(user_id, users_data)
    losing_admin = current["role"] == "admin" and new_role != "admin"
    losing_active = current["role"] == "admin" and current["status"] == "active" and new_status != "active"
    deleting_admin = current["role"] == "admin" and new_status == "deleted"

    if (losing_admin or losing_active or deleting_admin) and not admin_users:
        raise AppError("At least one active admin user must remain in the system.")


def validate_employee_form(form):
    employee_code = (form.get("employee_code") or "").strip()
    full_name = (form.get("full_name") or "").strip()
    department = (form.get("department") or "").strip()
    status = (form.get("status") or "active").strip().lower()

    if not employee_code:
        raise AppError("Employee code is required.")
    if not full_name:
        raise AppError("Employee name is required.")
    if not department:
        raise AppError("Department is required.")
    if status not in {"active", "inactive"}:
        raise AppError("Status must be either active or inactive.")

    return {
        "employee_code": employee_code,
        "full_name": full_name,
        "department": department,
        "status": status,
    }


def validate_user_form(form, is_new):
    username = (form.get("username") or "").strip()
    full_name = (form.get("full_name") or "").strip()
    role = (form.get("role") or "user").strip().lower()
    status = (form.get("status") or "active").strip().lower()
    password = form.get("password") or ""
    confirm_password = form.get("confirm_password") or ""
    security_question = (form.get("security_question") or "").strip()
    security_answer = (form.get("security_answer") or "").strip()

    if not username:
        raise AppError("Username is required.")
    if not full_name:
        raise AppError("Full name is required.")
    if role not in {"admin", "user"}:
        raise AppError("Role must be admin or user.")
    if status not in {"active", "inactive"}:
        raise AppError("Status must be active or inactive.")
    if not security_question:
        raise AppError("Security question is required.")
    if is_new:
        if not security_answer:
            raise AppError("Security answer is required.")
        validate_new_password(password, confirm_password)

    return {
        "username": username,
        "full_name": full_name,
        "role": role,
        "status": status,
        "password": password,
        "confirm_password": confirm_password,
        "security_question": security_question,
        "security_answer": security_answer,
    }


def validate_new_password(password, confirm_password):
    if len(password) < 6:
        raise AppError("Password must be at least 6 characters long.")
    if password != confirm_password:
        raise AppError("Password and confirm password do not match.")


def parse_cooldown_seconds(raw_value):
    value = (raw_value or "").strip()
    if not value:
        raise AppError("Cooldown time is required.")
    try:
        cooldown = int(value)
    except ValueError as exc:
        raise AppError("Cooldown time must be a whole number of seconds.") from exc
    if cooldown < 0:
        raise AppError("Cooldown time cannot be negative.")
    return cooldown


def parse_required_int(raw_value, label):
    try:
        return int(raw_value)
    except (TypeError, ValueError) as exc:
        raise AppError(f"{label} value is invalid.") from exc


def decode_data_url_image(data_url):
    if not data_url:
        raise AppError("Camera image is missing. Please capture a frame and try again.")
    if "," not in data_url:
        raise AppError("Invalid camera image received.")

    _, encoded = data_url.split(",", 1)
    try:
        image_bytes = base64.b64decode(encoded)
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except (binascii.Error, OSError, ValueError) as exc:
        raise AppError("Invalid camera image received.") from exc
    return np.array(image)


def extract_single_face_encoding(frame):
    face_locations = face_recognition.face_locations(frame)
    if not face_locations:
        raise AppError("No face detected. Center your face in the camera and try again.")
    if len(face_locations) > 1:
        raise AppError("Multiple faces detected. Make sure only one face is visible.")

    encodings = face_recognition.face_encodings(frame, face_locations)
    if not encodings:
        raise AppError("Face detected, but encoding failed. Please try again.")
    return encodings[0]


def serialize_encoding(encoding):
    return json.dumps([float(value) for value in encoding])


def deserialize_encoding(serialized_encoding):
    return np.array(json.loads(serialized_encoding), dtype=np.float64)


def save_employee_face(employee_id, employee_code, frame):
    FACE_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    file_name = f"{employee_id}_{sanitize_filename(employee_code)}.jpg"
    relative_path = Path("data") / "employee_faces" / file_name
    absolute_path = BASE_DIR / relative_path

    image_bgr = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
    cv2.imwrite(str(absolute_path), image_bgr)
    return relative_path.as_posix()


def sanitize_filename(value):
    return "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in value)


def parse_iso_datetime(value):
    return datetime.fromisoformat(value)


def get_latest_scan(employee_id, attendance_rows):
    rows = [row for row in attendance_rows if row["employee_id"] == employee_id]
    if not rows:
        return None
    return max(rows, key=lambda row: row["attended_at"])


def build_today_summaries(employees, attendance_rows, decisions):
    return build_day_summaries(date.today().isoformat(), employees, attendance_rows, decisions)


def build_day_summaries(attended_on, employees, attendance_rows, decisions):
    summaries = []
    for employee in employees:
        decision = find_decision(employee["id"], attended_on, decisions)
        summaries.append(build_employee_day_summary(employee, attended_on, attendance_rows, decision))
    summaries.sort(key=lambda item: item["full_name"].lower())
    return summaries


def build_employee_daily_summaries(employee_id, attendance_rows, decisions):
    days = {row["attended_on"] for row in attendance_rows if row["employee_id"] == employee_id}
    days.update(
        decision["attended_on"] for decision in decisions if decision["employee_id"] == employee_id
    )

    employee = get_employee_or_404(employee_id)
    summaries = []
    for attended_on in sorted(days, reverse=True):
        decision = find_decision(employee_id, attended_on, decisions)
        summaries.append(build_employee_day_summary(employee, attended_on, attendance_rows, decision))
    return summaries


def build_employee_day_summary(employee, attended_on, attendance_rows, decision=None):
    timestamps = sorted(
        row["attended_at"]
        for row in attendance_rows
        if row["employee_id"] == employee["id"] and row["attended_on"] == attended_on
    )

    first_time = timestamps[0] if timestamps else ""
    last_time = timestamps[-1] if timestamps else ""
    duration_seconds = 0
    if len(timestamps) >= 2:
        duration_seconds = int(
            (parse_iso_datetime(last_time) - parse_iso_datetime(first_time)).total_seconds()
        )
        duration_seconds = max(duration_seconds, 0)

    if duration_seconds >= PRESENT_SECONDS:
        auto_status = "present"
    elif duration_seconds >= HALFDAY_SECONDS:
        auto_status = "halfday"
    else:
        auto_status = "review"

    ot_seconds = max(duration_seconds - PRESENT_SECONDS, 0)
    final_status = decision["final_status"] if decision else ""
    resolved_status = final_status or ("pending" if auto_status == "review" else auto_status)

    return {
        "employee_id": employee["id"],
        "full_name": employee["full_name"],
        "employee_code": employee["employee_code"],
        "department": employee["department"],
        "attended_on": attended_on,
        "timestamps": timestamps,
        "timestamps_text": ", ".join(format_time_only(value) for value in timestamps) if timestamps else "No scans",
        "first_time": format_time_only(first_time) if first_time else "-",
        "last_time": format_time_only(last_time) if last_time else "-",
        "duration_seconds": duration_seconds,
        "duration_text": format_seconds(duration_seconds),
        "auto_status": auto_status,
        "resolved_status": resolved_status,
        "ot_seconds": ot_seconds,
        "ot_text": format_seconds(ot_seconds),
        "decision_notes": decision["notes"] if decision else "",
        "decision_by": decision["decided_by"] if decision else "",
        "review_needed": auto_status == "review",
    }


def format_seconds(total_seconds):
    total_seconds = int(total_seconds or 0)
    if total_seconds <= 0:
        return "0m"
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    parts = []
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    if seconds and not hours:
        parts.append(f"{seconds}s")
    return " ".join(parts) if parts else "0m"


def format_time_only(iso_value):
    return parse_iso_datetime(iso_value).strftime("%H:%M:%S")
